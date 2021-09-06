from web3.main import Web3
import openpyxl  # Thu vien excel]
import os  # system
try:
    print('Waiting...')

    infura_url = 'https://mainnet.infura.io/v3/6f1c1b42d64a48998728d9d6781306f2'
    w3 = Web3(Web3.HTTPProvider(infura_url))
    # config
    if os.path.isfile('./Result.xlsx'):
        wb = openpyxl.load_workbook('Result.xlsx')
        sheet = wb.active
        print('Result.xlsx exist')
        print('Total wallet: ', sheet.max_row - 1)
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active   
        sheet['A1'] = 'STT'
        sheet['B1'] = 'Address'
        sheet['C1'] = 'Private Key'    
        print('Result.xlsx does not exist')
        print('Created Resuilt.xlsx')
        sheet.column_dimensions['B'].width = 50
        sheet.column_dimensions['C'].width = 80
    row = sheet.max_row
        
    while True:
        print('Submit new wallet = ', end='')
        try:
            Total = int(input())
            if Total<0:
                raise
        except Exception as e:            
            print(e)
            print('Submit agains')
            continue
        break
    os.system('cls')
    #create wallet and write in excel
    for i in range(1, Total + 1, 1):
        print('Creating ' + str(i) + 'st Wallet')
        acct = w3.eth.account.create('')
        sheet.cell(i + row, 1, i + row - 1)
        sheet.cell(i + row, 2, acct.address)
        sheet.cell(i + row, 3, str(acct.key))
    wb.save('Result.xlsx')
# LocalAccount
    print('Created ', Total, ' Wallet')
    print('Total Wallet: ', sheet.max_row-1)
    print('Done, please check Result.xlsx')
    print('My phone: 0822579013')
except Exception as e:
    print('Error!!!')
    print(e)
print('press Enter to continue...')
input() 
os.system('cls')